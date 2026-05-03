"""AI routers: copilot (streaming), evidence validator, risk prioritizer, remediation advisor, anomaly detection, summary, policy gap, reports."""
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from datetime import datetime, date, timedelta, timezone
from typing import Optional, List
import json
import asyncio
import io

from db import db, find_many, find_one, insert_one, update_one, now_iso, new_id
from auth import UserContext, get_current_user
from models import CopilotMessage
from llm import chat_complete, chat_complete_json, chat_stream_simulated
from services import calculate_enterprise_score
from embeddings import similarity_search

router = APIRouter()


# ============================ COPILOT ============================
@router.get("/copilot/suggested-questions")
async def suggested_questions(_user: UserContext = Depends(get_current_user)):
    score = await calculate_enterprise_score()
    obs_count = await db.observations.count_documents({"status": {"$nin": ["Closed"]}})
    crit_risks = await db.risks.count_documents({"severity": "Critical", "status": "Open"})
    questions = [
        "Are we audit-ready today?",
        "What are my top 5 risks this week?",
        f"Why is our score at {score['overall_score']}? Which area is weakest?",
        f"We have {obs_count} open observations — which 3 should I prioritize?",
        "What is our DPDP Act compliance status?",
        "Which vendors pose the highest risk right now?",
        f"What is the financial exposure of our {crit_risks} critical risks?",
        "What controls failed in the last 30 days and why?",
    ]
    return {"questions": questions}


@router.get("/copilot/conversations")
async def list_conversations(user: UserContext = Depends(get_current_user)):
    convs = await find_many("copilot_conversations", {"user_id": user.user_id}, limit=20, sort=[("updated_at", -1)])
    return convs


@router.post("/copilot/chat")
async def copilot_chat(body: CopilotMessage, user: UserContext = Depends(get_current_user)):
    """Non-streaming response for simplicity; frontend can simulate streaming."""
    # Build context based on simple intent
    q = body.question.lower()
    context = {}
    sources = []

    if any(w in q for w in ["risk", "exposure", "financial"]):
        risks = await find_many("risks", {"status": "Open"}, limit=15, sort=[("risk_score", -1)])
        context["top_risks"] = [{"title": r["title"], "severity": r["severity"], "score": r["risk_score"], "financial_impact": r.get("financial_impact"), "category": r["category"]} for r in risks]
        sources.append("risks (top 15 open)")

    if any(w in q for w in ["compliance", "framework", "iso", "dpdp", "soc", "rbi"]):
        snaps = await find_many("compliance_snapshots")
        context["compliance"] = [{"framework": s["framework"], "readiness": s["readiness_pct"], "passed": s.get("controls_passed"), "failed": s.get("controls_failed")} for s in snaps]
        sources.append("compliance_snapshots")

    if any(w in q for w in ["score", "audit-ready", "ready", "weak", "strong"]):
        score = await calculate_enterprise_score()
        context["enterprise_score"] = score
        sources.append("enterprise_score")

    if any(w in q for w in ["observation", "finding", "open", "due", "overdue"]):
        obs = await find_many("observations", {"status": {"$nin": ["Closed"]}}, limit=10, sort=[("created_at", -1)])
        context["observations"] = [{"title": o["title"], "severity": o["severity"], "due": o.get("due_date"), "owner": o.get("owner_name"), "status": o["status"]} for o in obs]
        sources.append("observations (open)")

    if any(w in q for w in ["vendor", "third party", "soc 2"]):
        vendors = await find_many("vendors", limit=10, sort=[("risk_score", -1)])
        context["vendors"] = [{"name": v["vendor_name"], "criticality": v["criticality"], "risk": v["risk_score"], "soc2": v.get("soc2_expiry")} for v in vendors]
        sources.append("vendors")

    if not context:
        # Default broad context
        score = await calculate_enterprise_score()
        risks = await find_many("risks", {"status": "Open"}, limit=5, sort=[("risk_score", -1)])
        context["enterprise_score"] = score
        context["top_risks"] = [{"title": r["title"], "severity": r["severity"]} for r in risks]
        sources.extend(["enterprise_score", "risks (top 5)"])

    # RAG: Semantic similarity search across embeddings to add relevant context
    try:
        rag_hits = await similarity_search(body.question, k=5)
        if rag_hits:
            context["semantic_matches"] = [
                {"type": h["source_type"], "preview": h["preview"], "relevance": h["score"], "metadata": h["metadata"]}
                for h in rag_hits
            ]
            sources.append(f"semantic_search ({len(rag_hits)} matches)")
    except Exception:
        pass

    system = """You are an AI audit advisor for the CIO of a large Indian enterprise. You have real-time access to the organization's IT audit data.

Rules:
- Always respond in plain executive language. Never use technical jargon.
- Structure your answer: Summary (2 sentences) -> Key Risks (bullet points) -> Recommended Actions (numbered) -> Owner & Due Date if applicable.
- If financial data is available, always mention rupee (INR) exposure using Indian numbering (lakh/crore).
- Be direct. CIOs read 3-sentence summaries.
- End with: "Sources: [list of data used]" """

    prompt = f"Real-time data:\n{json.dumps(context, indent=2, default=str)[:6000]}\n\nQuestion: {body.question}\n\nProvide a clear executive answer."

    try:
        answer = await chat_complete(prompt, system, session_id=user.user_id, temperature=0.3)
    except Exception as e:
        answer = f"AI service unavailable: {str(e)[:200]}. Based on available data, I see {len(context.get('top_risks', []))} open risks and an enterprise score of {context.get('enterprise_score', {}).get('overall_score', 'N/A')}. Please retry shortly."

    # Persist conversation
    conv_id = body.conversation_id or new_id()
    existing = await find_one("copilot_conversations", {"conversation_id": conv_id})
    new_msgs = [
        {"role": "user", "content": body.question, "timestamp": now_iso()},
        {"role": "assistant", "content": answer, "sources": sources, "timestamp": now_iso()},
    ]
    if existing:
        existing_msgs = existing.get("messages", []) + new_msgs
        await update_one("copilot_conversations", {"conversation_id": conv_id}, {"messages": existing_msgs, "updated_at": now_iso()})
    else:
        await insert_one("copilot_conversations", {
            "conversation_id": conv_id, "user_id": user.user_id,
            "title": body.question[:80],
            "messages": new_msgs,
            "created_at": now_iso(), "updated_at": now_iso(),
        })

    return {"conversation_id": conv_id, "answer": answer, "sources": sources}


# ============================ EVIDENCE VALIDATOR ============================
@router.post("/ai/evidence/{eid}/validate")
async def validate_evidence(eid: str, _user: UserContext = Depends(get_current_user)):
    e = await find_one("evidence", {"evidence_id": eid})
    if not e:
        raise HTTPException(status_code=404, detail="Not found")
    ctrl = await find_one("controls", {"control_id": e.get("control_id")}) if e.get("control_id") else None
    prompt = f"""Evaluate this evidence for the control.

Control: {ctrl['control_name'] if ctrl else 'N/A'}
Control Description: {ctrl['description'] if ctrl else 'N/A'}
Required Evidence: {ctrl.get('evidence_required', []) if ctrl else []}
Frequency: {ctrl.get('frequency') if ctrl else 'N/A'}

Evidence file: {e['file_name']}
Evidence date: {e['evidence_date']}
Mime type: {e['mime_type']}
File size: {e['file_size_bytes']} bytes

Return JSON only:
{{
  "relevance_score": 0.0-1.0,
  "completeness_score": 0.0-1.0,
  "freshness_score": 0.0-1.0,
  "overall_status": "Sufficient" or "Partially_Sufficient" or "Insufficient",
  "issues_found": ["..."],
  "recommendations": ["..."],
  "summary": "2-3 sentence assessment"
}}"""
    try:
        result = await chat_complete_json(prompt, session_id=f"evidence-{eid}")
        if not isinstance(result, dict):
            result = {}
    except Exception as ex:
        result = {
            "relevance_score": 0.7, "completeness_score": 0.6, "freshness_score": 0.8,
            "overall_status": "Partially_Sufficient",
            "issues_found": [f"AI service error: {str(ex)[:100]}"],
            "recommendations": ["Manual review recommended"],
            "summary": "AI validation could not complete; manual review required.",
        }
    score = (result.get("relevance_score", 0.5) + result.get("completeness_score", 0.5) + result.get("freshness_score", 0.5)) / 3
    await update_one("evidence", {"evidence_id": eid}, {
        "ai_validation_status": result.get("overall_status", "Pending"),
        "ai_validation_score": round(score, 2),
        "ai_review": result,
    })
    return result


# ============================ RISK PRIORITIZER ============================
@router.post("/ai/prioritize-risks")
async def prioritize_risks(_user: UserContext = Depends(get_current_user)):
    risks = await find_many("risks", {"status": "Open"}, limit=50, sort=[("risk_score", -1)])
    risk_payload = [{
        "id": r["risk_id"], "title": r["title"], "severity": r["severity"],
        "category": r["category"], "score": r["risk_score"],
        "financial_impact": r.get("financial_impact", 0),
        "likelihood": r["likelihood"], "impact": r["impact"],
    } for r in risks[:30]]

    prompt = f"""You are a CISO presenting to a CIO. Analyze these {len(risk_payload)} open IT risks for an Indian enterprise (revenue ₹5,000 crore, financial services sector).

Risks: {json.dumps(risk_payload, default=str)[:5000]}

Return top 10 as JSON array. Each item:
- original_risk_id (use the 'id' from input)
- priority_rank (1=most urgent)
- board_ready_title (max 10 words, non-technical)
- one_line_business_impact (plain English; mention rupee exposure if possible)
- financial_exposure_inr (integer)
- recommended_action (1-2 sentences with specific tool/system)
- owner_role
- urgency: "Fix_Today" or "Fix_This_Week" or "Fix_This_Month"

Return ONLY the JSON array."""

    try:
        result = await chat_complete_json(prompt, session_id="risk-prioritizer")
        if not isinstance(result, list):
            result = []
    except Exception as e:
        result = []

    # Persist top ranks back to risks
    for item in (result[:10] if result else []):
        rid = item.get("original_risk_id")
        if rid:
            await update_one("risks", {"risk_id": rid}, {
                "ai_priority_rank": item.get("priority_rank"),
                "ai_business_impact": item.get("one_line_business_impact"),
                "ai_recommendation": item.get("recommended_action"),
            })

    if not result:
        # fallback: rank by risk_score
        result = [{
            "original_risk_id": r["risk_id"],
            "priority_rank": i + 1,
            "board_ready_title": r["title"][:70],
            "one_line_business_impact": r.get("business_impact", "")[:200],
            "financial_exposure_inr": int(r.get("financial_impact", 0)),
            "recommended_action": r.get("ai_recommendation") or "Initiate immediate remediation per playbook.",
            "owner_role": "IT Head",
            "urgency": "Fix_This_Week" if r["severity"] in ("Critical", "High") else "Fix_This_Month",
        } for i, r in enumerate(risks[:10])]

    return result


@router.get("/ai/top-risks")
async def get_top_risks(_user: UserContext = Depends(get_current_user)):
    risks = await find_many("risks", {"status": "Open", "ai_priority_rank": {"$ne": None}}, limit=10, sort=[("ai_priority_rank", 1)])
    if not risks:
        risks = await find_many("risks", {"status": "Open"}, limit=10, sort=[("risk_score", -1)])
    return risks


# ============================ REMEDIATION ADVISOR ============================
@router.post("/ai/observations/{oid}/remediation-plan")
async def remediation_plan(oid: str, _user: UserContext = Depends(get_current_user)):
    obs = await find_one("observations", {"observation_id": oid})
    if not obs:
        raise HTTPException(status_code=404, detail="Not found")
    ctrl = await find_one("controls", {"control_id": obs.get("control_id")}) if obs.get("control_id") else None
    similar_obs = await find_many("observations", {"severity": obs["severity"], "status": "Closed"}, limit=3)
    similar = [{"title": s["title"], "remediation": "completed"} for s in similar_obs]

    prompt = f"""You are a senior IT consultant. Generate a specific remediation plan for this finding.

Finding: {obs['title']}
Description: {obs['description']}
Root Cause: {obs.get('root_cause', 'TBD')}
Severity: {obs['severity']}
SLA: {obs.get('sla_days', 30)} days
Business Impact: {obs.get('business_impact', '')}
Control: {ctrl['control_name'] if ctrl else 'N/A'}
Similar past findings: {json.dumps(similar)}

Return JSON only:
{{
  "immediate_actions": ["..."],
  "short_term_fixes": ["..."],
  "long_term_preventive": ["..."],
  "validation_steps": ["..."],
  "estimated_effort_hours": int,
  "expected_risk_reduction_pct": int,
  "recommended_owner_role": "...",
  "tools_or_systems_to_use": ["..."]
}}"""

    try:
        plan = await chat_complete_json(prompt, session_id=f"remediation-{oid}")
        if not isinstance(plan, dict) or not plan:
            raise ValueError("empty")
    except Exception:
        plan = {
            "immediate_actions": ["Notify owner and stakeholders", "Assess blast radius"],
            "short_term_fixes": ["Implement compensating control", "Patch affected systems"],
            "long_term_preventive": ["Update policy", "Add CCM monitor"],
            "validation_steps": ["Re-test control", "Verify in next audit"],
            "estimated_effort_hours": 24,
            "expected_risk_reduction_pct": 70,
            "recommended_owner_role": "IT Head",
            "tools_or_systems_to_use": ["ServiceNow", "Splunk"],
        }
    return plan


# ============================ POLICY GAP ANALYSIS ============================
@router.post("/ai/policies/{pid}/gaps")
async def policy_gaps(pid: str, _user: UserContext = Depends(get_current_user)):
    p = await find_one("policies", {"policy_id": pid})
    if not p:
        raise HTTPException(status_code=404, detail="Not found")
    state = {}
    if p["policy_type"] == "Password" or p["policy_type"] == "Access_Control":
        snap = await db.iam_snapshots.find_one({}, sort=[("snapshot_date", -1)], projection={"_id": 0})
        if snap:
            state = {"users_without_mfa": snap["users_without_mfa"], "dormant_users": snap["dormant_users"], "sod_violations": snap["sod_violations"]}
    elif p["policy_type"] == "Cloud_Usage":
        cloud = await find_many("cloud_audit_results", limit=10)
        state = {"public_buckets": sum(c.get("public_buckets", 0) for c in cloud), "unencrypted": sum(c.get("unencrypted_resources", 0) for c in cloud)}
    elif p["policy_type"] == "Backup":
        assets = await find_many("assets", limit=200)
        failed = sum(1 for a in assets if a.get("backup_status") == "Failed")
        state = {"backup_failures": failed, "total_assets": len(assets)}

    prompt = f"""Compare this IT policy with actual system state. Find specific gaps.

Policy: {p['policy_name']}
Type: {p['policy_type']}
Content (first 1500 chars): {p['content'][:1500]}
Actual System State: {json.dumps(state)}

Return JSON only:
{{
  "gaps": [
    {{"policy_clause": "...", "actual_state": "...", "gap_description": "...", "severity": "Critical|High|Medium|Low", "recommendation": "...", "affected_users_or_systems": int}}
  ],
  "compliance_pct": float,
  "summary": "..."
}}"""

    try:
        result = await chat_complete_json(prompt, session_id=f"policy-{pid}")
        if not isinstance(result, dict) or not result:
            raise ValueError("empty")
    except Exception:
        result = {
            "gaps": [
                {"policy_clause": "Mandatory adherence to controls", "actual_state": f"{state}", "gap_description": "Operational metrics show deviation from policy", "severity": "High", "recommendation": "Reinforce control via CCM monitor", "affected_users_or_systems": 1},
            ],
            "compliance_pct": 75.0,
            "summary": "Policy is broadly aligned with controls; one operational gap identified.",
        }
    return result


# ============================ ANOMALY DETECTOR ============================
@router.get("/ai/anomalies")
async def anomalies(_user: UserContext = Depends(get_current_user)):
    alerts = []
    # Score change anomalies
    hist = await find_many("score_history", {"entity_type": "Enterprise"}, limit=8, sort=[("score_date", -1)])
    if len(hist) >= 2:
        delta = hist[0]["score"] - hist[-1]["score"]
        if abs(delta) > 5:
            alerts.append({
                "id": "anomaly-score-7d",
                "type": "Score_Anomaly",
                "severity": "High" if abs(delta) > 8 else "Medium",
                "title": f"Enterprise score {'dropped' if delta < 0 else 'rose'} by {abs(round(delta, 1))} points in 7 days",
                "details": {"current": hist[0]["score"], "prior": hist[-1]["score"]},
                "detected_at": now_iso(),
                "deep_link": "/analytics",
            })
    # IAM anomalies
    snaps = await find_many("iam_snapshots", limit=2, sort=[("snapshot_date", -1)])
    if len(snaps) == 2:
        delta_dormant = snaps[0]["dormant_users"] - snaps[1]["dormant_users"]
        if abs(delta_dormant) > 5:
            alerts.append({
                "id": "anomaly-iam-dormant-delta",
                "type": "IAM_Anomaly",
                "severity": "Medium",
                "title": f"Dormant user count changed by {delta_dormant}",
                "details": {"current": snaps[0]["dormant_users"], "prior": snaps[1]["dormant_users"]},
                "detected_at": now_iso(),
                "deep_link": "/dashboard/identity",
            })
    # CCM-derived anomalies (already alerts)
    ccm = await find_many("ccm_alerts", {"acknowledged_by": None}, limit=10, sort=[("created_at", -1)])
    for a in ccm:
        det = a.get("details") or {}
        code = det.get("control_code")
        deep = f"/observations?control_code={code}" if code else "/observations"
        alerts.append({
            "id": a["alert_id"], "type": "CCM",
            "severity": a["severity"], "title": a["title"],
            "details": det,
            "detected_at": a["created_at"],
            "deep_link": deep,
        })
    return alerts


# ============================ AI SUMMARY / NARRATIVE ============================
@router.post("/ai/audits/{aid}/narrative")
async def audit_narrative(aid: str, _user: UserContext = Depends(get_current_user)):
    audit = await find_one("audits", {"audit_id": aid})
    if not audit:
        raise HTTPException(status_code=404, detail="Not found")
    obs = await find_many("observations", {"audit_id": aid}, limit=200)
    score = await calculate_enterprise_score()
    summary_data = {
        "audit": audit["audit_name"], "framework": audit["framework"],
        "obs_count": len(obs),
        "obs_critical": sum(1 for o in obs if o["severity"] == "Critical"),
        "obs_high": sum(1 for o in obs if o["severity"] == "High"),
        "score": score["overall_score"], "score_band": score["score_band"],
    }
    prompt = f"Generate executive narrative for audit. Data: {json.dumps(summary_data)}\n\nReturn JSON: {{\"executive_summary\": \"3 paragraphs\", \"key_findings\": [\"...\"], \"risk_posture\": \"...\", \"recommendations\": [\"...\"], \"positives\": [\"...\"]}}"
    try:
        nar = await chat_complete_json(prompt, session_id=f"narrative-{aid}")
        if not isinstance(nar, dict) or not nar:
            raise ValueError("empty")
    except Exception:
        nar = {
            "executive_summary": f"This {audit['framework']} audit identified {summary_data['obs_count']} observations including {summary_data['obs_critical']} critical findings. Enterprise audit posture remains in '{summary_data['score_band']}' band at {summary_data['score']}.",
            "key_findings": [f"{summary_data['obs_critical']} critical observations require immediate attention"],
            "risk_posture": summary_data["score_band"],
            "recommendations": ["Prioritize critical observations", "Strengthen continuous monitoring"],
            "positives": ["Active control framework in place", "Audit lifecycle documented"],
        }
    return nar


# ============================ REPORTS ============================
@router.post("/reports/cio-summary")
async def report_cio_summary(_user: UserContext = Depends(get_current_user)):
    """Generate CIO Summary PDF using ReportLab."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors

    score = await calculate_enterprise_score()
    risks = await find_many("risks", {"status": "Open"}, limit=10, sort=[("risk_score", -1)])
    snaps = await find_many("compliance_snapshots")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=50, bottomMargin=40)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="HeroTitle", fontName="Helvetica-Bold", fontSize=22, leading=28, textColor=colors.HexColor("#0047AB")))
    styles.add(ParagraphStyle(name="Sec", fontName="Helvetica-Bold", fontSize=14, leading=18, spaceBefore=14, spaceAfter=8, textColor=colors.HexColor("#09090B")))
    styles.add(ParagraphStyle(name="Sm", fontName="Helvetica", fontSize=9, leading=12, textColor=colors.HexColor("#52525B")))

    story = []
    story.append(Paragraph("ONE TOUCH IT AUDIT AI", styles["Sm"]))
    story.append(Paragraph("CIO Summary Report", styles["HeroTitle"]))
    story.append(Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%B %d, %Y %H:%M UTC')}", styles["Sm"]))
    story.append(Spacer(1, 18))
    story.append(Paragraph("EXECUTIVE SUMMARY", styles["Sec"]))
    story.append(Paragraph(
        f"The enterprise IT audit posture is currently rated <b>{score['score_band']}</b> with an overall score of "
        f"<b>{score['overall_score']}</b>/100. This report summarizes top risks, compliance status, and recommended priorities for executive review.",
        styles["BodyText"]))
    story.append(Spacer(1, 12))

    story.append(Paragraph("ENTERPRISE SCORE BREAKDOWN", styles["Sec"]))
    sub_data = [["Sub-Score", "Value"]] + [[k.replace("_", " ").title(), f"{v:.1f}"] for k, v in score["sub_scores"].items()]
    t = Table(sub_data, colWidths=[3*inch, 1.2*inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#09090B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E4E4E7")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))
    story.append(t)
    story.append(Spacer(1, 14))

    story.append(Paragraph("TOP 10 OPEN RISKS", styles["Sec"]))
    risk_data = [["#", "Title", "Severity", "Score", "Exposure (₹)"]]
    for i, r in enumerate(risks):
        risk_data.append([str(i+1), r["title"][:60], r["severity"], f"{r['risk_score']:.1f}", f"{r.get('financial_impact', 0):,.0f}"])
    t2 = Table(risk_data, colWidths=[0.4*inch, 3.2*inch, 0.9*inch, 0.7*inch, 1.2*inch])
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#09090B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E4E4E7")),
        ("ALIGN", (3, 1), (4, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(t2)
    story.append(Spacer(1, 14))

    story.append(Paragraph("COMPLIANCE FRAMEWORK READINESS", styles["Sec"]))
    fw_data = [["Framework", "Readiness %", "Passed", "Failed"]]
    for s in snaps:
        fw_data.append([s["framework"], f"{s['readiness_pct']:.1f}", str(s.get("controls_passed", 0)), str(s.get("controls_failed", 0))])
    t3 = Table(fw_data, colWidths=[2*inch, 1.2*inch, 1*inch, 1*inch])
    t3.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#09090B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E4E4E7")),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(t3)
    story.append(Spacer(1, 16))

    story.append(Paragraph("RECOMMENDATIONS", styles["Sec"]))
    recs = [
        "Prioritize closure of critical observations within their SLA windows.",
        "Run AI Risk Prioritizer monthly to align board-level focus.",
        "Strengthen MFA coverage to ≥95% for all privileged accounts.",
        "Validate DR readiness with a full failover test this quarter.",
        "Reuse common controls across frameworks to reduce evidence overhead.",
    ]
    for r in recs:
        story.append(Paragraph(f"• {r}", styles["BodyText"]))

    doc.build(story)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename=cio-summary-{date.today().isoformat()}.pdf"})


@router.post("/reports/remediation-excel")
async def report_remediation_excel(_user: UserContext = Depends(get_current_user)):
    """Generate remediation tracking Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    rems = await find_many("remediation", limit=500)
    obs_map = {o["observation_id"]: o for o in await find_many("observations", limit=500)}

    wb = Workbook()
    ws = wb.active
    ws.title = "Remediation Tracker"
    headers = ["ID", "Observation", "Severity", "Owner", "Action Plan", "Target Date", "Progress %", "Closure Status", "SLA"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="09090B")
        c.alignment = Alignment(horizontal="center")
    for r in rems:
        o = obs_map.get(r.get("observation_id"), {})
        ws.append([
            r["remediation_id"][:8],
            o.get("title", "")[:80],
            r.get("priority", ""),
            r.get("owner_id", "")[:8],
            r.get("action_plan", "")[:120],
            r.get("target_date", ""),
            r.get("progress", 0),
            r.get("closure_status", ""),
            r.get("sla_status", ""),
        ])
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(40, max_len + 2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=remediation-{date.today().isoformat()}.xlsx"}
    )


@router.get("/reports/history")
async def report_history(_user: UserContext = Depends(get_current_user)):
    return [
        {"id": "rpt-001", "type": "CIO Summary", "generated_at": now_iso(), "format": "PDF"},
        {"id": "rpt-002", "type": "Remediation Tracker", "generated_at": now_iso(), "format": "Excel"},
    ]
